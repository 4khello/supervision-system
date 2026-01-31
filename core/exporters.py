from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Count, Q

from core.models import Research, Supervisor, ResearchSupervision


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _auto_fit(ws):
    for col_cells in ws.columns:
        max_len = 0
        col = col_cells[0].column
        for c in col_cells:
            val = c.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)


def _apply_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.rightToLeft = True


def _status_filter_q(sf: str):
    """
    sf:
      active: الحاليين فقط (يستبعد ناقش/فصل/إلغاء)
      discussed: ناقش/انتهى
      active_discussed: الحاليين + ناقشوا (يستبعد فصل/إلغاء)
      dismissed: فصل
      all: الكل
    """
    sf = (sf or "active").strip().lower()

    if sf == "discussed":
        return Q(status=Research.Status.DISCUSSED), "discussed"

    if sf == "dismissed":
        return Q(status=Research.Status.DISMISSED), "dismissed"

    if sf == "all":
        return Q(), "all"

    if sf == "active_discussed":
        # الحاليين + ناقشوا = كل شيء ما عدا (فصل/إلغاء)
        return ~Q(status__in=[Research.Status.DISMISSED, Research.Status.CANCELLED]), "active_discussed"

    # default active: الحاليين فقط (يستبعد ناقش/فصل/إلغاء)
    excluded_for_active_only = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]
    return ~Q(status__in=excluded_for_active_only), "active"


def build_export_workbook(q: str = "", supervisor_id: Optional[str] = None, sf: str = "active") -> Workbook:
    wb = Workbook()

    status_filter, sf = _status_filter_q(sf)

    # -----------------------------------
    # شيت 1: Researches
    # -----------------------------------
    ws1 = wb.active
    ws1.title = "Researches"

    headers = [
        "Research ID",
        "اسم الباحث",
        "النوع",
        "المرحلة",
        "الحالة",
        "عنوان البحث",
        "المشرفون",
        "قسم المشرف (إن وجد)",
        "تاريخ التصدير",
        "فلتر التصدير (sf)",
    ]
    ws1.append(headers)

    researches = Research.objects.all().filter(status_filter)

    if q:
        researches = researches.filter(Q(researcher_name__icontains=q) | Q(title__icontains=q))

    if supervisor_id:
        researches = researches.filter(researchsupervision__supervisor_id=int(supervisor_id))

    researches = researches.distinct().prefetch_related("researchsupervision_set__supervisor__department")

    export_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    for r in researches:
        links = list(
            ResearchSupervision.objects.filter(research=r)
            .select_related("supervisor__department")
            .order_by("supervisor__name")
        )
        sup_names = " | ".join([l.supervisor.name for l in links])
        sup_depts = " | ".join([(l.supervisor.department.name if l.supervisor.department else "—") for l in links])

        ws1.append([
            r.id,
            r.researcher_name,
            r.get_researcher_type_display(),
            r.get_degree_display(),
            r.get_status_display(),
            r.title,
            sup_names,
            sup_depts,
            export_time,
            sf,
        ])

    _apply_header(ws1)
    _auto_fit(ws1)

    # -----------------------------------
    # شيت 2: Supervisors Summary
    # (✅ بدون عمود "إجمالي الروابط")
    # -----------------------------------
    ws2 = wb.create_sheet("Supervisors Summary")
    ws2.append([
        "Supervisor ID",
        "اسم المشرف",
        "القسم",
        "ماجستير (باحث)",
        "دكتوراه (باحث)",
        "إجمالي الباحثين",
        "المعيدين",
        "فلتر التصدير (sf)",
    ])

    # علشان summary يراعي sf: لازم يكون في filter على researches__status
    if sf == "discussed":
        status_q = Q(researches__status=Research.Status.DISCUSSED)
    elif sf == "dismissed":
        status_q = Q(researches__status=Research.Status.DISMISSED)
    elif sf == "all":
        status_q = Q()
    elif sf == "active_discussed":
        status_q = ~Q(researches__status__in=[Research.Status.DISMISSED, Research.Status.CANCELLED])
    else:
        status_q = ~Q(researches__status__in=[Research.Status.DISCUSSED, Research.Status.DISMISSED, Research.Status.CANCELLED])

    supervisors = (
        Supervisor.objects.filter(is_active=True)
        .select_related("department")
        .annotate(
            ma_count=Count(
                "researches",
                filter=(
                    Q(researches__researcher_type=Research.ResearcherType.RESEARCHER)
                    & Q(researches__degree=Research.Degree.MA)
                    & status_q
                ),
                distinct=True,
            ),
            phd_count=Count(
                "researches",
                filter=(
                    Q(researches__researcher_type=Research.ResearcherType.RESEARCHER)
                    & Q(researches__degree=Research.Degree.PHD)
                    & status_q
                ),
                distinct=True,
            ),
            researchers_total=Count(
                "researches",
                filter=(Q(researches__researcher_type=Research.ResearcherType.RESEARCHER) & status_q),
                distinct=True,
            ),
            assistants_count=Count(
                "researches",
                filter=(Q(researches__researcher_type=Research.ResearcherType.ASSISTANT) & status_q),
                distinct=True,
            ),
        )
        .order_by("-researchers_total", "name")
    )

    if q:
        supervisors = supervisors.filter(name__icontains=q)

    # لو supervisor_id موجود: نطلع صف واحد للمشرف ده
    if supervisor_id:
        supervisors = supervisors.filter(id=int(supervisor_id))

    for s in supervisors:
        ws2.append([
            s.id,
            s.name,
            s.department.name if s.department else "—",
            s.ma_count,
            s.phd_count,
            s.researchers_total,
            s.assistants_count,
            sf,
        ])

    _apply_header(ws2)
    _auto_fit(ws2)

    # -----------------------------------
    # شيت 3: Stats (طبق sf على الإحصائيات)
    # -----------------------------------
    ws3 = wb.create_sheet("Stats")
    ws3.append(["البند", "القيمة", "الإجمالي", "فلتر التصدير (sf)"])

    base = Research.objects.filter(status_filter)

    by_degree = base.values("degree").annotate(total=Count("id")).order_by("degree")
    by_status = base.values("status").annotate(total=Count("id")).order_by("-total")
    by_type = base.values("researcher_type").annotate(total=Count("id")).order_by("researcher_type")

    ws3.append(["حسب الدرجة", "", "", sf])
    for x in by_degree:
        ws3.append(["درجة", x["degree"], x["total"], sf])

    ws3.append(["", "", "", ""])
    ws3.append(["حسب الحالة", "", "", sf])
    for x in by_status:
        ws3.append(["حالة", x["status"], x["total"], sf])

    ws3.append(["", "", "", ""])
    ws3.append(["حسب النوع", "", "", sf])
    for x in by_type:
        ws3.append(["نوع", x["researcher_type"], x["total"], sf])

    _apply_header(ws3)
    _auto_fit(ws3)

    return wb