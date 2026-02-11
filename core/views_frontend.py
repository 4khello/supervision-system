from __future__ import annotations

import json
import re
from datetime import datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from core.exporters import build_export_workbook
from core.models import (
    Department,
    DepartmentUser,
    Research,
    ResearchFeePayment,
    ResearchSupervision,
    Supervisor,
)

# ============================================================
# Helpers
# ============================================================

def get_user_department(user):
    """Return Department for department users; superusers see all."""
    if getattr(user, "is_superuser", False):
        return None
    try:
        return user.department_user.department
    except Exception:
        return None


def can_edit(user) -> bool:
    """Only superusers can edit/add/delete/export/upload."""
    return bool(user and getattr(user, "is_superuser", False))


def get_research_scope_qs(user):
    """
    Research queryset scope for the current user.

    - Superuser: all Research
    - Department user: researches linked to supervisors of their department
      (NOT relying on Research.department)
    """
    qs = Research.objects.all()
    dept = get_user_department(user)
    if not dept:
        return qs
    dept_supers = Supervisor.objects.filter(department=dept, is_active=True)
    return qs.filter(researchsupervision__supervisor__in=dept_supers).distinct()


def get_supervisor_scope_qs(user):
    """Supervisor queryset scope for the current user."""
    qs = Supervisor.objects.filter(is_active=True).select_related("department")
    dept = get_user_department(user)
    if not dept:
        return qs
    return qs.filter(department=dept)


# ============================================================
# Index: open login مباشرة
# ============================================================

def index(request):
    # ✅ لما تشغل البروجيكت: يفتح login لو مش عامل تسجيل دخول
    if request.user.is_authenticated:
        return redirect("home")
    return redirect("login")


# ============================================================
# Auth (frontend)
# ============================================================

def login_view(request):
    if request.user.is_authenticated:
        return render(request, "registration/login.html")


    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()
        user = authenticate(request, username=username, password=password)
        if user is None:
            messages.error(request, "بيانات الدخول غير صحيحة")
            return render(request, "frontend/login.html", {"username": username})

        login(request, user)
        return redirect("home")

    return render(request, "frontend/login.html")


def register_view(request):
    # موجودة كاسم قديم لو محتاجها لاحقاً
    return render(request, "frontend/register.html")


@login_required
def logout_view(request):
    """
    Logout على GET و POST (علشان ما يطلعش 405)
    """
    logout(request)
    return redirect("login")


# ✅ Aliases علشان أي Template/URL قديم مايتكسرش
frontend_logout = logout_view
logout_user = logout_view   # لو template بيستخدم logout_user


# ============================================================
# Home
# ============================================================

@login_required
def home(request):
    """
    الصفحة الرئيسية + اختيار القسم (للأدمن فقط)، ولـ Dept user يتم إجباره على قسمه.
    الإحصائيات العلوية حسب نطاق اليوزر (Superuser=كل البيانات / Dept user=قسمه عبر مشرفين القسم)
    """
    dept_restriction = get_user_department(request.user)

    # ✅ Dept user: اجبر dept_id على قسمه
    if dept_restriction:
        if request.GET.get("dept_id") and str(request.GET.get("dept_id")) != str(dept_restriction.id):
            return redirect(f"/home/?dept_id={dept_restriction.id}")
        if not request.GET.get("dept_id"):
            return redirect(f"/home/?dept_id={dept_restriction.id}")

    excluded_statuses = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]

    # ✅ الإحصائيات العلوية حسب نطاق اليوزر
    qs_all = get_research_scope_qs(request.user)

    ma_current = (
        qs_all.filter(degree=Research.Degree.MA, researcher_type=Research.ResearcherType.RESEARCHER)
        .exclude(status__in=excluded_statuses)
        .count()
    )
    phd_current = (
        qs_all.filter(degree=Research.Degree.PHD, researcher_type=Research.ResearcherType.RESEARCHER)
        .exclude(status__in=excluded_statuses)
        .count()
    )
    discussed_count = qs_all.filter(status=Research.Status.DISCUSSED).count()
    dismissed_count = qs_all.filter(status=Research.Status.DISMISSED).count()

    # ✅ الأقسام
    departments_with_supervisors = (
        Supervisor.objects.filter(is_active=True, department__isnull=False)
        .values_list("department_id", flat=True)
        .distinct()
    )
    if dept_restriction:
        departments = Department.objects.filter(id=dept_restriction.id)
    else:
        departments = Department.objects.filter(id__in=departments_with_supervisors).order_by("name")

    dept_id = request.GET.get("dept_id")
    selected_dept = None
    dept_stats = None
    dept_supervisors = []

    if dept_id:
        selected_dept = get_object_or_404(Department, id=int(dept_id))

        dept_supervisors_qs = Supervisor.objects.filter(department=selected_dept, is_active=True)

        base_researches = Research.objects.filter(
            researchsupervision__supervisor__in=dept_supervisors_qs
        ).distinct()

        base_active = base_researches.exclude(status__in=excluded_statuses)

        dept_stats = {
            "total": base_active.filter(researcher_type=Research.ResearcherType.RESEARCHER).count(),
            "phd": base_active.filter(degree=Research.Degree.PHD, researcher_type=Research.ResearcherType.RESEARCHER).count(),
            "ma": base_active.filter(degree=Research.Degree.MA, researcher_type=Research.ResearcherType.RESEARCHER).count(),
            "assistants": base_active.filter(researcher_type=Research.ResearcherType.ASSISTANT).count(),
        }

        for supervisor in dept_supervisors_qs:
            supervisor_researches = base_active.filter(researchsupervision__supervisor=supervisor).distinct()

            ma_count = supervisor_researches.filter(
                degree=Research.Degree.MA, researcher_type=Research.ResearcherType.RESEARCHER
            ).count()
            phd_count = supervisor_researches.filter(
                degree=Research.Degree.PHD, researcher_type=Research.ResearcherType.RESEARCHER
            ).count()
            assistants_count = supervisor_researches.filter(
                researcher_type=Research.ResearcherType.ASSISTANT
            ).count()
            total_count = ma_count + phd_count

            supervisor.ma_count = ma_count
            supervisor.phd_count = phd_count
            supervisor.assistants_count = assistants_count
            supervisor.total_count = total_count

            dept_supervisors.append(supervisor)

        dept_supervisors = sorted(dept_supervisors, key=lambda x: x.total_count, reverse=True)

    # Chart data
    chart_data = {"labels": [], "phd": [], "ma": [], "assistants": []}
    for supervisor in dept_supervisors:
        short_name = " ".join((supervisor.name or "").split()[:2]) or (supervisor.name or "")
        chart_data["labels"].append(short_name)
        chart_data["phd"].append(getattr(supervisor, "phd_count", 0))
        chart_data["ma"].append(getattr(supervisor, "ma_count", 0))
        chart_data["assistants"].append(getattr(supervisor, "assistants_count", 0))

    return render(
        request,
        "frontend/home.html",
        {
            "ma_current": ma_current,
            "phd_current": phd_current,
            "discussed_count": discussed_count,
            "dismissed_count": dismissed_count,
            "departments": departments,
            "selected_dept": selected_dept,
            "selected_dept_id": selected_dept.id if selected_dept else None,
            "dept_stats": dept_stats,
            "dept_supervisors": dept_supervisors,
            "chart_data": chart_data,
            "dept_restriction": dept_restriction,
            "is_admin": request.user.is_superuser,
        },
    )


@login_required
def home_stat_details(request, stat_type):
    excluded_statuses = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]

    qs_api = get_research_scope_qs(request.user)

    if stat_type == "ma_current":
        researches = qs_api.filter(
            degree=Research.Degree.MA, researcher_type=Research.ResearcherType.RESEARCHER
        ).exclude(status__in=excluded_statuses)
        title = "ماجستير (الحاليين)"
    elif stat_type == "phd_current":
        researches = qs_api.filter(
            degree=Research.Degree.PHD, researcher_type=Research.ResearcherType.RESEARCHER
        ).exclude(status__in=excluded_statuses)
        title = "دكتوراه (الحاليين)"
    elif stat_type == "discussed":
        researches = qs_api.filter(status=Research.Status.DISCUSSED)
        title = "ناقشوا"
    elif stat_type == "dismissed":
        researches = qs_api.filter(status=Research.Status.DISMISSED)
        title = "مفصولين"
    else:
        return JsonResponse({"error": "Invalid stat type"}, status=400)

    researches = researches.select_related("department").prefetch_related("researchsupervision_set__supervisor")[:200]

    data = []
    for r in researches:
        supervisors = [f"د. {link.supervisor.name}" for link in r.researchsupervision_set.all()]
        data.append(
            {
                "id": r.id,
                "name": r.researcher_name,
                "degree": r.get_degree_display(),
                "supervisors": ", ".join(supervisors),
                "status": r.get_status_display(),
            }
        )

    return JsonResponse({"title": title, "count": len(data), "data": data})


# ✅ Alias عشان NoReverseMatch اللي ظهر عندك: home_department_stats
@login_required
def home_department_stats(request):
    # ببساطة يفتح صفحة احصائيات الأقسام
    return redirect("department_stats")


# ============================================================
# Dashboard (اختياري)
# ============================================================

@login_required
def dashboard(request):
    # لو عندك dashboard template قديم - خليه redirect عشان ما يطلب ملف ناقص
    return redirect("home")


# ============================================================
# Supervisors
# ============================================================

@login_required
def supervisors_page(request):
    q = (request.GET.get("q") or "").strip()
    dept_id = request.GET.get("dept_id")

    dept_restriction = get_user_department(request.user)
    if dept_restriction:
        dept_id = str(dept_restriction.id)

    excluded_for_active = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]
    status_q = ~Q(researches__status__in=excluded_for_active)

    supervisors = (
        get_supervisor_scope_qs(request.user)
        .annotate(
            ma_count=Count(
                "researches",
                filter=(
                    Q(researches__researcher_type=Research.ResearcherType.RESEARCHER)
                    & Q(researches__degree=Research.Degree.MA)
                    & status_q
                ),
                distinct=True,
            ),
            phd_count=Count(
                "researches",
                filter=(
                    Q(researches__researcher_type=Research.ResearcherType.RESEARCHER)
                    & Q(researches__degree=Research.Degree.PHD)
                    & status_q
                ),
                distinct=True,
            ),
            researchers_total=Count(
                "researches",
                filter=(Q(researches__researcher_type=Research.ResearcherType.RESEARCHER) & status_q),
                distinct=True,
            ),
            assistants_count=Count(
                "researches",
                filter=(Q(researches__researcher_type=Research.ResearcherType.ASSISTANT) & status_q),
                distinct=True,
            ),
            total_links=Count("researchsupervision", distinct=True),
        )
        .order_by("-researchers_total", "name")
    )

    if q:
        supervisors = supervisors.filter(name__icontains=q)

    if dept_id:
        supervisors = supervisors.filter(department_id=int(dept_id))

    if dept_restriction:
        all_departments = Department.objects.filter(id=dept_restriction.id)
    else:
        all_departments = Department.objects.filter(
            id__in=Supervisor.objects.filter(is_active=True, department__isnull=False)
            .values_list("department_id", flat=True)
            .distinct()
        ).order_by("name")

    return render(
        request,
        "frontend/supervisors.html",
        {
            "supervisors": supervisors,
            "q": q,
            "dept_id": dept_id,
            "all_departments": all_departments,
            "is_admin": request.user.is_superuser,
        },
    )


@login_required
def supervisor_detail(request, pk: int):
    supervisor = get_object_or_404(Supervisor.objects.select_related("department"), pk=pk)

    dept_restriction = get_user_department(request.user)
    if dept_restriction and supervisor.department_id != dept_restriction.id:
        messages.error(request, "غير مصرح لك بعرض هذا المشرف")
        return redirect("home")

    sf = (request.GET.get("sf") or "active").strip().lower()
    excluded_for_active_only = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]

    if sf == "discussed":
        status_filter = Q(status=Research.Status.DISCUSSED)
    elif sf == "dismissed":
        status_filter = Q(status=Research.Status.DISMISSED)
    elif sf == "all":
        status_filter = Q()
    elif sf == "active_discussed":
        status_filter = ~Q(status__in=[Research.Status.DISMISSED, Research.Status.CANCELLED])
    else:
        sf = "active"
        status_filter = ~Q(status__in=excluded_for_active_only)

    researches = (
        Research.objects.filter(researchsupervision__supervisor=supervisor)
        .filter(status_filter)
        .prefetch_related("researchsupervision_set__supervisor", "department")
        .order_by("-degree", "researcher_name")
        .distinct()
    )

    items = []
    for r in researches:
        links = list(
            ResearchSupervision.objects.filter(research=r)
            .select_related("supervisor")
            .order_by("supervisor__name")
        )
        co_supers = [l.supervisor for l in links if l.supervisor_id != supervisor.id]
        items.append({"research": r, "co_supervisors": co_supers})

    ma_count = sum(
        1
        for x in items
        if x["research"].researcher_type == Research.ResearcherType.RESEARCHER
        and x["research"].degree == Research.Degree.MA
    )
    phd_count = sum(
        1
        for x in items
        if x["research"].researcher_type == Research.ResearcherType.RESEARCHER
        and x["research"].degree == Research.Degree.PHD
    )
    researchers_total = ma_count + phd_count
    assistants_only = sum(1 for x in items if x["research"].researcher_type == Research.ResearcherType.ASSISTANT)

    return render(
        request,
        "frontend/supervisor_detail.html",
        {
            "supervisor": supervisor,
            "items": items,
            "ma_count": ma_count,
            "phd_count": phd_count,
            "researchers_total": researchers_total,
            "assistants_only": assistants_only,
            "sf": sf,
            "is_admin": request.user.is_superuser,
        },
    )


# ============================================================
# Researchers
# ============================================================

@login_required
def researchers_page(request):
    q = (request.GET.get("q") or "").strip()
    sf = (request.GET.get("sf") or "active").strip().lower()
    date_from = (request.GET.get("date_from") or "").strip() or None
    date_to = (request.GET.get("date_to") or "").strip() or None

    excluded_for_active = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]

    if sf == "discussed":
        status_filter = Q(status=Research.Status.DISCUSSED)
    elif sf == "dismissed":
        status_filter = Q(status=Research.Status.DISMISSED)
    elif sf == "all":
        status_filter = Q()
    elif sf == "active_discussed":
        status_filter = ~Q(status__in=[Research.Status.DISMISSED, Research.Status.CANCELLED])
    else:
        sf = "active"
        status_filter = ~Q(status__in=excluded_for_active)

    qs = get_research_scope_qs(request.user).filter(researcher_type=Research.ResearcherType.RESEARCHER)

    researches = (
        qs.filter(status_filter)
        .prefetch_related("researchsupervision_set__supervisor", "department")
        .order_by("-id")
    )

    if q:
        researches = researches.filter(Q(researcher_name__icontains=q) | Q(title__icontains=q))
    if date_from:
        researches = researches.filter(registration_date__gte=date_from)
    if date_to:
        researches = researches.filter(registration_date__lte=date_to)

    return render(
        request,
        "frontend/researchers.html",
        {
            "researches": researches,
            "q": q,
            "sf": sf,
            "date_from": date_from,
            "date_to": date_to,
            "is_admin": request.user.is_superuser,
            "dept_restriction": get_user_department(request.user),
        },
    )


@login_required
def research_detail(request, pk):
    qs = (
        get_research_scope_qs(request.user)
        .select_related("department")
        .prefetch_related("researchsupervision_set__supervisor", "fee_payments")
    )
    research = get_object_or_404(qs, pk=pk)

    supervisors = [link.supervisor for link in research.researchsupervision_set.all()]
    current_year = timezone.localdate().year

    ResearchFeePayment.objects.get_or_create(research=research, year=current_year, defaults={"is_paid": False})
    payments = list(research.fee_payments.all())

    return render(
        request,
        "frontend/research_detail.html",
        {
            "research": research,
            "supervisors": supervisors,
            "payments": payments,
            "current_year": current_year,
            "is_admin": request.user.is_superuser,
        },
    )


# ============================================================
# Department Stats
# ============================================================

@login_required
def department_stats(request):
    dept_restriction = get_user_department(request.user)
    dept_id = request.GET.get("dept_id")

    if dept_restriction:
        dept_id = str(dept_restriction.id)

    sf = (request.GET.get("sf") or "active").strip().lower()
    excluded_for_active = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]

    if sf == "discussed":
        status_filter = Q(status=Research.Status.DISCUSSED)
    elif sf == "dismissed":
        status_filter = Q(status=Research.Status.DISMISSED)
    elif sf == "all":
        status_filter = Q()
    elif sf == "active_discussed":
        status_filter = ~Q(status__in=[Research.Status.DISMISSED, Research.Status.CANCELLED])
    else:
        sf = "active"
        status_filter = ~Q(status__in=excluded_for_active)

    if dept_restriction:
        departments_with_supervisors = [dept_restriction.id]
    else:
        departments_with_supervisors = (
            Supervisor.objects.filter(is_active=True, department__isnull=False)
            .values_list("department_id", flat=True)
            .distinct()
        )

    all_departments = Department.objects.filter(id__in=departments_with_supervisors)

    departments = []
    for dept in all_departments:
        dept_supervisors = Supervisor.objects.filter(department=dept, is_active=True)
        dept_researches = (
            Research.objects.filter(researchsupervision__supervisor__in=dept_supervisors)
            .filter(status_filter)
            .distinct()
        )
        total = dept_researches.count()
        if total > 0:
            dept.total = total
            dept.phd = dept_researches.filter(degree=Research.Degree.PHD).count()
            dept.ma = dept_researches.filter(degree=Research.Degree.MA).count()
            dept.researchers = dept_researches.filter(researcher_type=Research.ResearcherType.RESEARCHER).count()
            dept.assistants = dept_researches.filter(researcher_type=Research.ResearcherType.ASSISTANT).count()
            departments.append(dept)

    departments = sorted(departments, key=lambda x: x.total, reverse=True)

    selected_dept = None
    dept_researches = []

    if dept_id:
        selected_dept = get_object_or_404(Department, id=int(dept_id))
        dept_supervisors = Supervisor.objects.filter(department=selected_dept, is_active=True)
        dept_researches = (
            Research.objects.filter(researchsupervision__supervisor__in=dept_supervisors)
            .filter(status_filter)
            .distinct()
            .prefetch_related("researchsupervision_set__supervisor")
        )

    return render(
        request,
        "frontend/department_stats.html",
        {
            "departments": departments,
            "selected_dept": selected_dept,
            "dept_researches": dept_researches,
            "sf": sf,
            "is_admin": request.user.is_superuser,
        },
    )


# ============================================================
# Exports (admin only)
# ============================================================

@login_required
def export_excel(request):
    if not can_edit(request.user):
        messages.error(request, "غير مصرح لك بالتصدير (الأدمن فقط).")
        return redirect("home")

    q = (request.GET.get("q") or "").strip()
    supervisor_id = request.GET.get("supervisor_id")
    sf = (request.GET.get("sf") or "active").strip().lower()

    wb = build_export_workbook(q=q, supervisor_id=supervisor_id, sf=sf)

    sf_slug = {
        "active": "active",
        "discussed": "discussed",
        "active_discussed": "active+discussed",
        "dismissed": "dismissed",
        "all": "all",
    }.get(sf, sf)

    def _safe_ascii_filename(text: str) -> str:
        text = (text or "").strip()
        text = re.sub(r"\s+", "_", text)
        text = re.sub(r"[^A-Za-z0-9_\-]+", "", text)
        return text[:40] or "Supervisor"

    date_part = datetime.now().strftime("%Y-%m-%d")

    if supervisor_id:
        sup = Supervisor.objects.filter(id=int(supervisor_id)).only("name").first()
        sup_part = _safe_ascii_filename(sup.name if sup else "Supervisor")
        filename_ascii = f"{sup_part}__{sf_slug}__{date_part}.xlsx"
    else:
        filename_ascii = f"Supervisors__{sf_slug}__{date_part}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename_ascii}"'
    wb.save(response)
    return response


@login_required
def export_department_excel(request):
    if not can_edit(request.user):
        messages.error(request, "غير مصرح لك بالتصدير (الأدمن فقط).")
        return redirect("home")

    dept_id = request.GET.get("dept_id")
    sf = (request.GET.get("sf") or "active").strip().lower()

    if not dept_id:
        return HttpResponse("يجب اختيار قسم", status=400)

    dept = get_object_or_404(Department, id=int(dept_id))

    excluded_for_active = [
        Research.Status.DISCUSSED,
        Research.Status.DISMISSED,
        Research.Status.CANCELLED,
    ]
    if sf == "discussed":
        status_filter = Q(status=Research.Status.DISCUSSED)
    elif sf == "dismissed":
        status_filter = Q(status=Research.Status.DISMISSED)
    elif sf == "all":
        status_filter = Q()
    elif sf == "active_discussed":
        status_filter = ~Q(status__in=[Research.Status.DISMISSED, Research.Status.CANCELLED])
    else:
        status_filter = ~Q(status__in=excluded_for_active)

    dept_supervisors = Supervisor.objects.filter(department=dept, is_active=True)
    researches = (
        Research.objects.filter(researchsupervision__supervisor__in=dept_supervisors)
        .filter(status_filter)
        .distinct()
        .prefetch_related("researchsupervision_set__supervisor")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{dept.name}"

    headers = ["#", "اسم الباحث", "النوع", "الدرجة", "عنوان الرسالة", "المشرفون", "الحالة"]
    ws.append(headers)

    for idx, research in enumerate(researches, start=1):
        supervisors = ", ".join([link.supervisor.name for link in research.researchsupervision_set.all()])
        researcher_type = "معيد" if research.researcher_type == Research.ResearcherType.ASSISTANT else "باحث"
        degree = "دكتوراه" if research.degree == Research.Degree.PHD else "ماجستير"
        ws.append([idx, research.researcher_name, researcher_type, degree, research.title or "", supervisors, research.get_status_display()])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="1a4f9c", end_color="1a4f9c", fill_type="solid")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    filter_name = {
        "active": "الحاليين",
        "discussed": "ناقشوا",
        "dismissed": "مفصولين",
        "all": "الكل",
        "active_discussed": "الحاليين+ناقشوا",
    }.get(sf, sf)
    filename = f"{dept.name}_{filter_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ============================================================
# Admin-only write operations (Dept users -> رسالة مش Error)
# ============================================================

@login_required
def add_researcher(request):
    if not can_edit(request.user):
        messages.error(request, "غير مسموح لك بالإضافة. (الأدمن فقط)")
        return redirect("researchers_page")
    
    # ✅ جلب البيانات المطلوبة للقوائم المنسدلة
    context = {
        "all_departments": Department.objects.all().order_by('name'),
        "all_supervisors": Supervisor.objects.filter(is_active=True).order_by('name'),
    }
    return render(request, "frontend/add_researcher.html", context)


@login_required
def edit_research(request, pk):
    if not can_edit(request.user):
        messages.error(request, "غير مسموح لك بالتعديل. (الأدمن فقط)")
        return redirect("research_detail", pk=pk)

    research = get_object_or_404(Research, pk=pk)
    
    # ✅ جلب البيانات للتعديل لضمان ظهور القوائم
    context = {
        "research": research,
        "all_departments": Department.objects.all().order_by('name'),
        "all_supervisors": Supervisor.objects.filter(is_active=True).order_by('name'),
        "current_supervisors": list(research.researchsupervision_set.values_list('supervisor_id', flat=True)),
        "payments": research.fee_payments.all().order_by('-year'),
    }
    return render(request, "frontend/edit_research.html", context)


@login_required
def delete_research(request, pk):
    if not can_edit(request.user):
        messages.error(request, "غير مسموح لك بالحذف. (الأدمن فقط)")
        return redirect("researchers_page")

    research = get_object_or_404(Research, pk=pk)
    name = research.researcher_name
    research.delete()
    messages.success(request, f"تم حذف الباحث: {name}")
    return redirect("researchers_page")


@login_required
def add_supervisor(request):
    if not can_edit(request.user):
        messages.error(request, "غير مسموح لك بالإضافة. (الأدمن فقط)")
        return redirect("supervisors_page")
    
    # ✅ جلب الأقسام علشان تظهر للمشرف الجديد
    context = {
        "all_departments": Department.objects.all().order_by('name'),
    }
    return render(request, "frontend/add_supervisor.html", context)


@login_required
def edit_supervisor(request, supervisor_id):
    if not can_edit(request.user):
        messages.error(request, "غير مسموح لك بالتعديل. (الأدمن فقط)")
        return redirect("supervisor_detail", pk=supervisor_id)

    supervisor = get_object_or_404(Supervisor, id=supervisor_id)
    
    if request.method == "POST":
        supervisor.name = request.POST.get("name")
        dept_id = request.POST.get("department_id")
        supervisor.department_id = int(dept_id) if dept_id else None
        supervisor.is_active = request.POST.get("is_active") == "on"
        supervisor.save()
        messages.success(request, f"تم تعديل بيانات د. {supervisor.name} بنجاح")
        return redirect("supervisor_detail", pk=supervisor.id)

    # ✅ التعديل هنا: نرسل 'departments' للـ template لكي تظهر في القائمة
    context = {
        "supervisor": supervisor,
        "departments": Department.objects.all().order_by('name'),
    }
    return render(request, "frontend/edit_supervisor.html", context)

@login_required
def delete_supervisor(request, pk):
    if not can_edit(request.user):
        messages.error(request, "غير مسموح لك بالحذف. (الأدمن فقط)")
        return redirect("supervisors_page")

    supervisor = get_object_or_404(Supervisor, pk=pk)
    name = supervisor.name
    ResearchSupervision.objects.filter(supervisor=supervisor).delete()
    supervisor.delete()
    messages.success(request, f"تم حذف المشرف: {name}")
    return redirect("supervisors_page")


# ============================================================
# Fees (API) - POST only + admin only
# ============================================================

@login_required
def toggle_fees_status(request, research_id, year):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not can_edit(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    research = get_object_or_404(Research, id=research_id)
    year = int(year)

    payment, _ = ResearchFeePayment.objects.get_or_create(
        research=research, year=year, defaults={"is_paid": False}
    )

    if payment.is_paid:
        payment.mark_unpaid()
        status = "unpaid"
    else:
        payment.mark_paid()
        status = "paid"

    return JsonResponse(
        {"success": True, "year": year, "status": status, "status_display": "دفع" if status == "paid" else "لم يدفع"}
    )


@login_required
def add_fees_year(request, research_id):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not can_edit(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    research = get_object_or_404(Research, id=research_id)
    year_raw = (request.POST.get("year") or "").strip()

    if not year_raw.isdigit():
        return JsonResponse({"error": "Year is required"}, status=400)

    year = int(year_raw)
    payment, created = ResearchFeePayment.objects.get_or_create(
        research=research, year=year, defaults={"is_paid": False}
    )

    return JsonResponse(
        {
            "success": True,
            "created": created,
            "year": year,
            "status": "paid" if payment.is_paid else "unpaid",
            "status_display": "دفع" if payment.is_paid else "لم يدفع",
        }
    )


@login_required
def delete_fees_year(request, research_id, year):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    if not can_edit(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    research = get_object_or_404(Research, id=research_id)
    deleted_count, _ = ResearchFeePayment.objects.filter(research=research, year=int(year)).delete()

    if deleted_count > 0:
        return JsonResponse({"success": True, "year": year, "message": f"تم حذف سنة {year} بنجاح"})
    return JsonResponse({"success": False, "error": "السنة غير موجودة"}, status=404)


# ============================================================
# Upload Researchers (admin only) - موجودة علشان urls مايتكسرش
# ============================================================

@login_required
def upload_researchers(request):
    if not can_edit(request.user):
        messages.error(request, "غير مصرح لك بالرفع (الأدمن فقط).")
        return redirect("home")
    return render(request, "frontend/upload_researchers.html")


# ============================================================
# Placeholder / Old names (علشان أي URLs قديمة ما تتكسرش)
# ============================================================

@login_required
def choose_page(request):
    return redirect("home")


@login_required
def stats_page(request):
    return redirect("home")


@login_required
def stat_details(request, stat_type):
    return JsonResponse({"error": "Not implemented"}, status=400)
